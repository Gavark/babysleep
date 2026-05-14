"""
Génère calculateur_sommeil_bebe.xlsx — à importer dans Google Sheets.

Trois feuilles : Calculateur, Historique, Paramètres.
Aucune macro (openpyxl ne sait pas créer une case à cocher Google Sheets).
L'utilisateur ajoute la case à cocher en B30 du Calculateur après import.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# --- Palette ---
FILL_YELLOW = PatternFill("solid", fgColor="FFF2CC")   # saisies
FILL_GREEN = PatternFill("solid", fgColor="C6E0B4")    # calculs-clés
FILL_HEADER = PatternFill("solid", fgColor="4472C4")   # section
FILL_ORANGE = PatternFill("solid", fgColor="FCE4D6")   # action
FILL_SUMMARY = PatternFill("solid", fgColor="DDEBF7")  # synthèse Historique
FILL_CALC_COL = PatternFill("solid", fgColor="E2EFDA") # colonnes calculées
FILL_ZEBRA = PatternFill("solid", fgColor="F7F7F7")    # alternance
FILL_THEAD = PatternFill("solid", fgColor="2F5496")    # en-tête tableau Historique

FONT_INPUT = Font(name="Calibri", size=11, bold=True, color="1F4E78")
FONT_CALC = Font(name="Calibri", size=11, bold=True, color="1F4E78")
FONT_HEADER = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
FONT_TITLE = Font(name="Calibri", size=18, bold=True, color="1F4E78")
FONT_SUBTITLE = Font(name="Calibri", size=11, italic=True, color="595959")
FONT_LABEL = Font(name="Calibri", size=11, color="262626")
FONT_THEAD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_TIME = "hh:mm"
FMT_DATE = "dd/mm/yyyy"
FMT_DUR = "[h]:mm"
FMT_ECART = '+0" min";-0" min";"à l\'heure"'


def style_input(cell, fmt=FMT_TIME):
    cell.fill = FILL_YELLOW
    cell.font = FONT_INPUT
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER
    cell.number_format = fmt


def style_calc(cell, fmt=FMT_TIME, *, key=False):
    cell.fill = FILL_GREEN if key else FILL_CALC_COL
    cell.font = FONT_CALC
    cell.alignment = ALIGN_CENTER
    cell.border = BORDER
    cell.number_format = fmt


def style_label(cell):
    cell.font = FONT_LABEL
    cell.alignment = ALIGN_LEFT


def section_header(ws, row, text, span=("A", "B")):
    ws.merge_cells(f"{span[0]}{row}:{span[1]}{row}")
    c = ws[f"{span[0]}{row}"]
    c.value = text
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


# ====== Workbook ======
wb = Workbook()

# Default sheet → Calculateur
ws_calc = wb.active
ws_calc.title = "Calculateur"

ws_hist = wb.create_sheet("Historique")
ws_param = wb.create_sheet("Paramètres")

# ====== Feuille Paramètres ======
param_headers = [
    "Tranche d'âge", "Âge min (mois)", "Âge max (mois)", "Nb siestes",
    "Fenêtre éveil (min)", "Fenêtre avant coucher (min)",
    "Sommeil nuit (h)", "Sommeil jour total (h)",
]
param_rows = [
    ("0-3 mois",   0,  3, 4, 75,   60,  9,   7),
    ("3-4 mois",   3,  4, 4, 90,   90, 10,   5),
    ("4-6 mois",   4,  6, 3, 120, 120, 11, 3.5),
    ("6-9 mois",   6,  9, 3, 165, 180, 11,   3),
    ("9-12 mois",  9, 12, 2, 210, 240, 11, 2.5),
    ("12-18 mois", 12, 18, 2, 270, 300, 11, 2.5),
    ("18-24 mois", 18, 24, 1, 330, 330, 11,   2),
    ("2-3 ans",    24, 36, 1, 360, 360, 11, 1.5),
]

for col, h in enumerate(param_headers, start=1):
    c = ws_param.cell(row=1, column=col, value=h)
    c.fill = FILL_THEAD
    c.font = FONT_THEAD
    c.alignment = ALIGN_CENTER
    c.border = BORDER

for r_idx, row in enumerate(param_rows, start=2):
    for c_idx, val in enumerate(row, start=1):
        c = ws_param.cell(row=r_idx, column=c_idx, value=val)
        c.border = BORDER
        c.alignment = ALIGN_CENTER if c_idx > 1 else ALIGN_LEFT
        if r_idx % 2 == 0:
            c.fill = FILL_ZEBRA

ws_param.column_dimensions["A"].width = 16
for col in "BCDEFGH":
    ws_param.column_dimensions[col].width = 18
ws_param.freeze_panes = "A2"

# ====== Feuille Calculateur ======
ws_calc.column_dimensions["A"].width = 48
ws_calc.column_dimensions["B"].width = 18
ws_calc.column_dimensions["C"].width = 4

# Titre + sous-titre (row 1-2 merged A:B)
ws_calc.merge_cells("A1:B1")
ws_calc["A1"] = "🍼 Calculateur de siestes et coucher"
ws_calc["A1"].font = FONT_TITLE
ws_calc["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws_calc.row_dimensions[1].height = 32

ws_calc.merge_cells("A2:B2")
ws_calc["A2"] = "Saisis uniquement les cases jaunes. Tout le reste se calcule automatiquement."
ws_calc["A2"].font = FONT_SUBTITLE
ws_calc["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws_calc.row_dimensions[2].height = 18

# Section: Profil
section_header(ws_calc, 4, "👶 Profil de l'enfant")

ws_calc["A5"] = "Âge en mois"
style_label(ws_calc["A5"])
ws_calc["B5"] = 6
style_input(ws_calc["B5"], fmt="0")

ws_calc["A6"] = "Heure de réveil souhaitée le matin"
style_label(ws_calc["A6"])
ws_calc["B6"] = 7 / 24  # 07:00 stored as fraction of day
style_input(ws_calc["B6"])

# Validation âge 0–36
dv_age = DataValidation(type="whole", operator="between", formula1=0, formula2=36,
                        allow_blank=False, showErrorMessage=True,
                        errorTitle="Âge invalide", error="L'âge doit être compris entre 0 et 36 mois.")
ws_param.add_data_validation  # noqa: unused, keep import explicit
ws_calc.add_data_validation(dv_age)
dv_age.add("B5")

# Section: Recommandations
section_header(ws_calc, 8, "📊 Recommandations pour cet âge")

recos = [
    (9,  "Tranche d'âge correspondante",
        "=INDEX(Paramètres!A2:A9, MATCH(B5, Paramètres!B2:B9, 1))", "@"),
    (10, "Nombre de siestes recommandé",
        "=VLOOKUP(B5, Paramètres!B2:H9, 3, TRUE)", "0"),
    (11, "Fenêtre d'éveil entre siestes (min)",
        "=VLOOKUP(B5, Paramètres!B2:H9, 4, TRUE)", "0"),
    (12, "Fenêtre d'éveil avant coucher (min)",
        "=VLOOKUP(B5, Paramètres!B2:H9, 5, TRUE)", "0"),
    (13, "Sommeil de nuit recommandé (h)",
        "=VLOOKUP(B5, Paramètres!B2:H9, 6, TRUE)", "0.0"),
    (14, "⭐ Coucher idéal (pour réveil souhaité)",
        "=MOD(B6 - B13/24, 1)", FMT_TIME),
]
for row, label, formula, fmt in recos:
    ws_calc.cell(row=row, column=1, value=label)
    style_label(ws_calc.cell(row=row, column=1))
    cell = ws_calc.cell(row=row, column=2, value=formula)
    style_calc(cell, fmt=fmt, key=(row == 14))

# Section: Plan de la journée
section_header(ws_calc, 16, "🗓️ Plan de la journée")

ws_calc["A17"] = "Réveil ce matin"
style_label(ws_calc["A17"])
style_input(ws_calc["B17"])

plan_pairs = [
    (18, "💤 Sieste 1 suggérée vers",                  "=MOD(B17 + B11/1440, 1)"),
    (19, "  → Fin réelle sieste 1",                    None),  # saisie
    (20, "💤 Sieste 2 suggérée vers",                  "=IF(ISBLANK(B19),\"\",MOD(B19 + B11/1440, 1))"),
    (21, "  → Fin réelle sieste 2",                    None),
    (22, "💤 Sieste 3 suggérée vers",                  "=IF(ISBLANK(B21),\"\",MOD(B21 + B11/1440, 1))"),
    (23, "  → Fin réelle sieste 3",                    None),
    (24, "💤 Sieste 4 suggérée vers (jeunes bébés)",   "=IF(ISBLANK(B23),\"\",MOD(B23 + B11/1440, 1))"),
    (25, "  → Fin réelle sieste 4",                    None),
]
for row, label, formula in plan_pairs:
    ws_calc.cell(row=row, column=1, value=label)
    style_label(ws_calc.cell(row=row, column=1))
    cell = ws_calc.cell(row=row, column=2)
    if formula is None:
        style_input(cell)
    else:
        cell.value = formula
        style_calc(cell, fmt=FMT_TIME)

# Row 27: coucher suggéré (vert key)
ws_calc["A27"] = "⭐ Coucher suggéré ce soir"
style_label(ws_calc["A27"])
ws_calc["B27"] = ("=IF(AND(ISBLANK(B19),ISBLANK(B21),ISBLANK(B23),ISBLANK(B25)),"
                  "B14,MOD(MAX(B17,B19,B21,B23,B25) + B12/1440, 1))")
style_calc(ws_calc["B27"], fmt=FMT_TIME, key=True)

# Row 28: coucher effectif (saisie)
ws_calc["A28"] = "🛏️ Coucher effectif (à remplir le soir)"
style_label(ws_calc["A28"])
style_input(ws_calc["B28"])

# Row 29: écart vs idéal
ws_calc["A29"] = "Écart vs coucher idéal"
style_label(ws_calc["A29"])
ws_calc["B29"] = "=IF(ISBLANK(B28), ROUND((B27-B14)*1440,0), ROUND((B28-B14)*1440,0))"
style_calc(ws_calc["B29"], fmt=FMT_ECART)

# Row 30: enregistrer (case à cocher manuelle)
ws_calc["A30"] = "💾 Enregistrer la journée → cocher la case"
ws_calc["A30"].font = Font(name="Calibri", size=11, bold=True, color="9C5700")
ws_calc["A30"].fill = FILL_ORANGE
ws_calc["A30"].alignment = ALIGN_LEFT
ws_calc["A30"].border = BORDER
ws_calc["B30"] = False  # checkbox sera ajoutée par l'utilisateur dans Sheets
ws_calc["B30"].fill = FILL_ORANGE
ws_calc["B30"].border = BORDER
ws_calc["B30"].alignment = ALIGN_CENTER
ws_calc["B30"].font = Font(name="Calibri", size=11, bold=True, color="9C5700")

# Section: Configuration initiale
section_header(ws_calc, 32, "⚙️ Configuration initiale (une fois)")
config_lines = [
    "1. Insertion → Case à cocher sur B30 (la cellule orange à droite).",
    "2. Extensions → Apps Script → coller le contenu de enregistrer_jour.gs → Enregistrer.",
    "3. Tester : remplir B17 (réveil), cocher B30 → la journée part dans l'Historique.",
    "4. Compatible mobile : ouvrir le fichier dans l'app Google Sheets et cocher B30 depuis ton tél.",
]
for i, txt in enumerate(config_lines, start=33):
    ws_calc.merge_cells(f"A{i}:B{i}")
    c = ws_calc[f"A{i}"]
    c.value = txt
    c.font = FONT_LABEL
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_calc.row_dimensions[i].height = 18

# Section: Rappels
section_header(ws_calc, 38, "💡 Rappels")
rappels = [
    "• Surveiller les signes de fatigue plutôt que l'horloge stricte.",
    "• Dernière sieste ≥ 2h-3h avant le coucher.",
    "• Bébé surfatigué = plus dur à endormir, mieux vaut coucher trop tôt.",
    "• Régularité > précision : viser le créneau, pas la minute exacte.",
    "• Si coucher suggéré très en retard sur l'idéal, raccourcir la dernière sieste demain.",
]
for i, txt in enumerate(rappels, start=39):
    ws_calc.merge_cells(f"A{i}:B{i}")
    c = ws_calc[f"A{i}"]
    c.value = txt
    c.font = FONT_LABEL
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_calc.row_dimensions[i].height = 18

ws_calc.sheet_view.showGridLines = False

# ====== Feuille Historique ======
ws_hist.sheet_view.showGridLines = False

# Synthèse (lignes 5–9)
ws_hist.merge_cells("A4:J4")
title = ws_hist["A4"]
title.value = "📈 Synthèse (sur les 100 jours du tableau)"
title.font = FONT_TITLE
title.alignment = Alignment(horizontal="left", vertical="center")
ws_hist.row_dimensions[4].height = 28

summary_rows = [
    (5, "Heure moyenne de réveil",        "=IFERROR(AVERAGE(B11:B110),\"\")",  FMT_TIME),
    (6, "Heure moyenne de coucher",       "=IFERROR(AVERAGE(G11:G110),\"\")",  FMT_TIME),
    (7, "Durée moyenne de nuit",          "=IFERROR(AVERAGE(I11:I110),\"\")",  FMT_DUR),
    (8, "Nombre moyen de siestes/jour",   "=IFERROR(AVERAGE(H11:H110),\"\")",  "0.0"),
    (9, "Nombre de jours enregistrés",    "=COUNT(A11:A110)",                    "0"),
]
for row, label, formula, fmt in summary_rows:
    ws_hist.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    lbl = ws_hist.cell(row=row, column=1, value=label)
    lbl.fill = FILL_SUMMARY
    lbl.font = Font(name="Calibri", size=11, bold=True, color="1F4E78")
    lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lbl.border = BORDER
    # Style merged cell row too
    for col in range(1, 3):
        ws_hist.cell(row=row, column=col).fill = FILL_SUMMARY
        ws_hist.cell(row=row, column=col).border = BORDER

    val = ws_hist.cell(row=row, column=3, value=formula)
    val.fill = FILL_GREEN
    val.font = Font(name="Calibri", size=11, bold=True, color="1F4E78")
    val.alignment = ALIGN_CENTER
    val.number_format = fmt
    val.border = BORDER

# Header row (ligne 10)
hist_headers = [
    ("Date", 12, FMT_DATE),
    ("Réveil", 10, FMT_TIME),
    ("Fin sieste 1", 12, FMT_TIME),
    ("Fin sieste 2", 12, FMT_TIME),
    ("Fin sieste 3", 12, FMT_TIME),
    ("Fin sieste 4", 12, FMT_TIME),
    ("Coucher", 10, FMT_TIME),
    ("Nb siestes", 11, "0"),
    ("Durée nuit préc.", 16, FMT_DUR),
    ("Notes", 30, "@"),
]
for col_idx, (name, width, _fmt) in enumerate(hist_headers, start=1):
    letter = get_column_letter(col_idx)
    ws_hist.column_dimensions[letter].width = width
    c = ws_hist.cell(row=10, column=col_idx, value=name)
    c.fill = FILL_THEAD
    c.font = FONT_THEAD
    c.alignment = ALIGN_CENTER
    c.border = BORDER
ws_hist.row_dimensions[10].height = 22

# Data rows 11–110
for r in range(11, 111):
    zebra = (r % 2 == 0)
    for col_idx, (_, _, fmt) in enumerate(hist_headers, start=1):
        c = ws_hist.cell(row=r, column=col_idx)
        c.number_format = fmt
        c.alignment = ALIGN_CENTER if col_idx != 10 else ALIGN_LEFT
        c.border = BORDER
        # Saisies (A–G + J)
        if col_idx in (1, 2, 3, 4, 5, 6, 7, 10):
            if zebra:
                c.fill = FILL_ZEBRA
        # Calculées (H, I)
        else:
            c.fill = FILL_CALC_COL

    ws_hist[f"H{r}"] = f"=IF(ISBLANK(A{r}),\"\",COUNT(C{r}:F{r}))"
    ws_hist[f"I{r}"] = f"=IF(AND(ISNUMBER(B{r}),ISNUMBER(G{r-1})),MOD(B{r}-G{r-1},1),\"\")"

ws_hist.freeze_panes = "B11"

# ====== Save ======
out = "calculateur_sommeil_bebe.xlsx"
wb.save(out)
print(f"OK {out} genere ({len(wb.sheetnames)} feuilles : {', '.join(wb.sheetnames)})")
